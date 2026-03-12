import re
import html
import uuid
import pandas as pd
from io import BytesIO
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from sqlalchemy.orm import Session

from config import settings
from models import MdExcelLog
from schemas_config import FIELD_MAPPING_CONFIG


class MarkdownService:
    @staticmethod
    def normalize(md: str) -> str:
        md = md.replace('\r\n', '\n').replace('\r', '\n')
        md = re.sub(r'(?m)^\s+', '', md)
        return md

    @staticmethod
    def clean_text(text: str) -> str:
        if not text: return ""
        text = html.unescape(text)
        text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
        text = re.sub(r'</p\s*>', '\n', text, flags=re.IGNORECASE)
        text = re.sub(r'<p\s*>', '', text, flags=re.IGNORECASE)
        text = re.sub(r'<[^>]+>', '', text)
        text = text.replace('**', '').replace('__', '').replace('`', '')
        return text.strip()

    def parse_to_dataframe(self, md_text: str) -> pd.DataFrame:
        """解析 Markdown 为 DataFrame"""
        text = self.normalize(md_text)
        
        # 正则匹配
        pattern = re.compile(
            r"(?:\n|^)(\|.+\|\n\|[\s\-:|]+\|\n(?:\|.*\|\n?)+)", 
            re.VERBOSE
        )
        match = pattern.search(text)
        if not match:
            raise ValueError("未找到有效的 Markdown 表格")

        table_md = match.group(1)
        lines = [line for line in table_md.splitlines() if line.strip()]
        
        headers = [self.clean_text(col.strip()) for col in lines[0].strip("|").split("|")]
        rows = []
        for line in lines[2:]:
            row = [self.clean_text(col.strip()) for col in line.strip("|").split("|")]
            rows.append(row)
            
        return pd.DataFrame(rows, columns=headers)

class ExcelService:
    def __init__(self):
        self.base_dir = Path(settings.FILE_STORAGE_PATH)
        self.base_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def generate_batch_id() -> str:
        date_str = datetime.now().strftime("%Y%m%d")
        random_str = uuid.uuid4().hex[:4]
        return f"{date_str}-{random_str}"

    def save_excel(self, df: pd.DataFrame) -> Path:
        filename = f"{uuid.uuid4().hex}.xlsx"
        filepath = self.base_dir / filename
        
        # 保存并设置样式
        df.to_excel(filepath, index=False, sheet_name="Sheet1")
        self._apply_style(filepath)
        return filepath

    def _apply_style(self, filepath: Path):
        """设置自动换行等样式"""
        wb = load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        wb.save(filepath)

    def get_bytes(self, df: pd.DataFrame) -> BytesIO:
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
        buffer.seek(0)
        return buffer



class DatabaseService:
    @staticmethod
    def save_conversion_records(db: Session, filename: str, batch_id: str, df: pd.DataFrame, model_class):
        """
        通用的数据库写入方法
        :param model_class: 动态生成的 ORM 类
        """
        if df.empty:
            return

        records = []
        
        # 遍历 DataFrame
        for _, row in df.iterrows():
            # 初始化字典，先填入系统字段
            data = {
                "filename": filename,
                "batch_id": batch_id,
                "status": "success"
            }
            
            # === 核心优化：自动映射业务字段 ===
            # 遍历配置文件，从 row 中查找对应的值
            for md_header, config in FIELD_MAPPING_CONFIG.items():
                db_col_name = config["db_column"]
                # 使用 get 防止列不存在报错，默认值为空字符串
                data[db_col_name] = row.get(md_header, "")
            
            # 创建 ORM 对象
            record = model_class(**data)
            records.append(record)
        
        try:
            db.add_all(records)
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"数据库写入失败: {e}")
            raise e